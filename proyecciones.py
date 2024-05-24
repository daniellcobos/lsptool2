import pandas as pd
from statsmodels.tsa.api import ExponentialSmoothing, SimpleExpSmoothing, Holt
import json
import plotly.express as px
import numpy as np
import plotly
import itertools
from openpyxl import load_workbook

def mape(y_true, y_pred):
    y_true, y_pred = np.array(y_true), np.array(y_pred)
    return np.mean(np.abs((y_true - y_pred) / y_true)) * 100

def proyeccioninicial(dt):
    # Convierte los periodos a fecha, proyeccion a 5 años
    pproyeccion = 72
    df0 = dt.copy(deep = True)
    df0['Periodo'] = pd.to_datetime(df0['Periodo'], format='%Y-%m-%d')
    tfechaarr = []
    tcrecarr = []
    tcolumnamearr = []
    figdict = {}
    # Determina las columnas y hace la proyeccion
    for (columnName, columnData) in df0.items():

        #Proyeccion inicial, todos tienen los mismos parametros
        ptrend = 'add'
        pseasonal = 'add'
        pdamped = True
        puse = True
        periodos = 12




        if (columnName != 'Periodo'):
            print(columnName)
            df10 = df0[df0[columnName] > 0]
            df10 = df10[['Periodo', columnName]]
            df10 = df10.set_index('Periodo')
            # fig = components.plot()
            # fig.set_size_inches(12, 6)
            # Proyeccion
            fit1 = ExponentialSmoothing(df10, seasonal_periods=periodos, trend=ptrend, seasonal=pseasonal,
                                        damped_trend=pdamped, use_boxcox=puse, initialization_method="estimated").fit()
            # print(fit1.summary())
            # Calcula el intervalo de confianza al 95% y lo salva
            xhat = fit1.forecast(steps=pproyeccion)

            z = 1.96
            sse = fit1.sse

            predint_xminus = xhat - z * np.sqrt(sse / len(df10))
            predint_xplus = xhat + z * np.sqrt(sse / len(df10))

            df = [predint_xminus, predint_xplus]
            df = pd.DataFrame(df)
            df = df.transpose()
            col0 = columnName + "_minimo"
            col1 = columnName + "_maximo"

            # Changing columns name with index number
            mapping = {df.columns[0]: col0, df.columns[1]: col1}
            df = df.rename(columns=mapping)

            # Changing columns name with index number
            mapping = {df.columns[0]: col0, df.columns[1]: col1}
            df = df.rename(columns=mapping)

            # Salva los intervalos


            # Intervalo de confianza por simulacion
            simulations = fit1.simulate(pproyeccion, repetitions=1, error="add", random_errors=None)
            # ax = df10.plot(figsize=(12, 5),marker="o",color="black",title="Proyeccion",)
            # ax.set_ylabel("Venta")
            # ax.set_xlabel("Periodo")
            # fit1.fittedvalues.plot(ax=ax, style="--", color="green")
            graphforecast = fit1.forecast(steps=pproyeccion)
            # graphforecast.rename("Proyeccion").plot(ax=ax, style="--", marker="o", color="green")

            # Grafica la serie de Ajuste y la añade a fit1
            df6 = fit1.forecast(pproyeccion)
            df7 = fit1.predict(0)

            df8 = [df7, df6]
            df10a = df10[columnName]
            df15 = pd.concat([df10a, df6])
            d = {'fecha': df15.index, 'crecimiento': df15[:].tolist()}
            tfechaarr.append(df15.index.tolist())
            tcrecarr.append(df15[:])
            tcolumnamearr.append([str(columnName) for x in df15])
            df16 = pd.DataFrame(data=d)
            fig2 = px.line(df16, x='fecha', y='crecimiento', title=columnName)

            fig2json = json.dumps(fig2, cls=plotly.utils.PlotlyJSONEncoder)
            figdict[columnName] = fig2json
            df1 = (df0[[columnName]])

            df = pd.DataFrame(df8)
            df = df.transpose()

            col0 = columnName + "_ajuste"
            col1 = columnName + "_proyeccion"

            # Changing columns name with index number
            mapping = {df.columns[0]: col0, df.columns[1]: col1}
            df = df.rename(columns=mapping)
            # Salva la proyeccion normal
            mfila = 'py_' + columnName + '.xlsx'
            df.to_excel(mfila, index_label="Periodo")

            # Calcula el Error por medio del MAPE
            df99 = [df7]
            yy = pd.DataFrame(df99)
            yy = yy.transpose()
            xx = pd.DataFrame(df10[[columnName]])
            print("MAPE " + columnName + " = " + str(mape(xx, yy)))



    tfechaarr = list(itertools.chain.from_iterable(tfechaarr))
    tcrecarr = list(itertools.chain.from_iterable(tcrecarr))
    tcolumnamearr = list(itertools.chain.from_iterable(tcolumnamearr))
    totaldata = {'fecha': tfechaarr, 'crecimiento': tcrecarr, 'categoria': tcolumnamearr}
    dftd = pd.DataFrame(data=totaldata)
    fig3 = px.line(dftd, x='fecha', y='crecimiento', title='Crecimientos por Categoria', color='categoria')
    figdict["final"] = json.dumps(fig3, cls=plotly.utils.PlotlyJSONEncoder)
    return figdict

def proyeccionmodify(dt,categoriar,ptrendr,pseasonalr,puser,pdampedr):
    # Convierte los periodos a fecha, proyeccion a 5 años
    pproyeccion = 72
    df0 = dt.copy(deep = True)
    df0['Periodo'] = pd.to_datetime(df0['Periodo'], format='%Y-%m-%d')
    tfechaarr = []
    tcrecarr = []
    tcolumnamearr = []
    figdict = {}
    # Determina las columnas y hace la proyeccion
    for (columnName, columnData) in df0.items():


        if columnName == categoriar:
            ptrend = ptrendr
            pseasonal = pseasonalr
            pdamped = pdampedr
            puse = puser
            periodos = 12





            print(columnName)
            df10 = df0[df0[columnName] > 0]
            df10 = df10[['Periodo', columnName]]
            df10 = df10.set_index('Periodo')
            # fig = components.plot()
            # fig.set_size_inches(12, 6)
            # Proyeccion
            fit1 = ExponentialSmoothing(df10, seasonal_periods=periodos, trend=ptrend, seasonal=pseasonal,
                                        damped_trend=pdamped, use_boxcox=puse, initialization_method="estimated").fit()
            # print(fit1.summary())
            # Calcula el intervalo de confianza al 95% y lo salva
            xhat = fit1.forecast(steps=pproyeccion)

            z = 1.96
            sse = fit1.sse

            predint_xminus = xhat - z * np.sqrt(sse / len(df10))
            predint_xplus = xhat + z * np.sqrt(sse / len(df10))

            df = [predint_xminus, predint_xplus]
            df = pd.DataFrame(df)
            df = df.transpose()
            col0 = columnName + "_minimo"
            col1 = columnName + "_maximo"

            # Changing columns name with index number
            mapping = {df.columns[0]: col0, df.columns[1]: col1}
            df = df.rename(columns=mapping)

            # Changing columns name with index number
            mapping = {df.columns[0]: col0, df.columns[1]: col1}
            df = df.rename(columns=mapping)

            # Salva los intervalos


            # Intervalo de confianza por simulacion
            simulations = fit1.simulate(pproyeccion, repetitions=1, error="add", random_errors=None)
            # ax = df10.plot(figsize=(12, 5),marker="o",color="black",title="Proyeccion",)
            # ax.set_ylabel("Venta")
            # ax.set_xlabel("Periodo")
            # fit1.fittedvalues.plot(ax=ax, style="--", color="green")
            graphforecast = fit1.forecast(steps=pproyeccion)
            # graphforecast.rename("Proyeccion").plot(ax=ax, style="--", marker="o", color="green")

            # Grafica la serie de Ajuste y la añade a fit1
            df6 = fit1.forecast(pproyeccion)
            df7 = fit1.predict(0)

            df8 = [df7, df6]
            df10a = df10[columnName]
            df15 = pd.concat([df10a, df6])
            d = {'fecha': df15.index, 'crecimiento': df15[:].tolist()}
            tfechaarr.append(df15.index.tolist())
            tcrecarr.append(df15[:])
            tcolumnamearr.append([str(columnName) for x in df15])
            df16 = pd.DataFrame(data=d)
            fig2 = px.line(df16, x='fecha', y='crecimiento', title=columnName)

            fig2json = json.dumps(fig2, cls=plotly.utils.PlotlyJSONEncoder)
            figdict[columnName] = fig2json
            df1 = (df0[[columnName]])

            df = pd.DataFrame(df8)
            df = df.transpose()

            col0 = columnName + "_ajuste"
            col1 = columnName + "_proyeccion"

            # Changing columns name with index number
            mapping = {df.columns[0]: col0, df.columns[1]: col1}
            df = df.rename(columns=mapping)
            # Salva la proyeccion normal
            mfila = 'py_' + columnName + '.xlsx'
            df.to_excel(mfila, index_label="Periodo")

            # Calcula el Error por medio del MAPE
            df99 = [df7]
            yy = pd.DataFrame(df99)
            yy = yy.transpose()
            xx = pd.DataFrame(df10[[columnName]])
            print("MAPE " + columnName + " = " + str(mape(xx, yy)))
    return figdict

def proyeccionParams(archivo,params):
    dt = pd.read_excel(archivo)
    # Convierte los periodos a fecha, proyeccion a 5 años
    pproyeccion = 72
    df0 = dt.copy(deep = True)
    df0['Periodo'] = pd.to_datetime(df0['Periodo'], format='%Y-%m-%d')
    tfechaarr = []
    tcrecarr = []
    tcolumnamearr = []
    figdict = {}

    # Determina las columnas y hace la proyeccion
    for (columnName, columnData) in df0.items():

        if columnName in params:
            #Proyeccion inicial, todos tienen los mismos parametros
            ptrend = params[columnName]["ptrend"]
            pseasonal =  params[columnName]["pseasonal"]
            pdamped = params[columnName]["pdamped"]
            puse =  float(params[columnName]["puse"])
            if puse >= 1:
                puse = True
            periodos = 12
            if ptrend:
                ptrend = "add"
            else:
                ptrend = "mul"
            if pseasonal:
                pseasonal = "add"
            else:
                pseasonal = "mul"
        else:
            ptrend = 'add'
            pseasonal = 'add'
            pdamped = True
            puse = True
            periodos = 12

        if (columnName != 'Periodo'):
            print(columnName)
            df10 = df0[df0[columnName] > 0]
            df10 = df10[['Periodo', columnName]]
            df10 = df10.set_index('Periodo')
            # fig = components.plot()
            # fig.set_size_inches(12, 6)
            # Proyeccion
            fit1 = ExponentialSmoothing(df10, seasonal_periods=periodos, trend=ptrend, seasonal=pseasonal,
                                        damped_trend=pdamped, use_boxcox=puse, initialization_method="estimated").fit()
            # print(fit1.summary())
            # Calcula el intervalo de confianza al 95% y lo salva
            xhat = fit1.forecast(steps=pproyeccion)

            z = 1.96
            sse = fit1.sse

            predint_xminus = xhat - z * np.sqrt(sse / len(df10))
            predint_xplus = xhat + z * np.sqrt(sse / len(df10))

            df = [predint_xminus, predint_xplus]
            df = pd.DataFrame(df)
            df = df.transpose()
            col0 = columnName + "_minimo"
            col1 = columnName + "_maximo"

            # Changing columns name with index number
            mapping = {df.columns[0]: col0, df.columns[1]: col1}
            df = df.rename(columns=mapping)

            # Changing columns name with index number
            mapping = {df.columns[0]: col0, df.columns[1]: col1}
            df = df.rename(columns=mapping)

            # Salva los intervalos


            # Intervalo de confianza por simulacion
            simulations = fit1.simulate(pproyeccion, repetitions=1, error="add", random_errors=None)
            # ax = df10.plot(figsize=(12, 5),marker="o",color="black",title="Proyeccion",)
            # ax.set_ylabel("Venta")
            # ax.set_xlabel("Periodo")
            # fit1.fittedvalues.plot(ax=ax, style="--", color="green")
            graphforecast = fit1.forecast(steps=pproyeccion)
            # graphforecast.rename("Proyeccion").plot(ax=ax, style="--", marker="o", color="green")

            # Grafica la serie de Ajuste y la añade a fit1
            df6 = fit1.forecast(pproyeccion)
            df7 = fit1.predict(0)

            df8 = [df7, df6]
            df10a = df10[columnName]
            df15 = pd.concat([df10a, df6])
            d = {'fecha': df15.index, 'crecimiento': df15[:].tolist()}
            tfechaarr.append(df15.index.tolist())
            tcrecarr.append(df15[:])
            tcolumnamearr.append([str(columnName) for x in df15])
            df16 = pd.DataFrame(data=d)
            fig2 = px.line(df16, x='fecha', y='crecimiento', title=columnName)

            fig2json = json.dumps(fig2, cls=plotly.utils.PlotlyJSONEncoder)
            figdict[columnName] = fig2json
            df1 = (df0[[columnName]])

            df = pd.DataFrame(df8)
            df = df.transpose()

            col0 = columnName + "_ajuste"
            col1 = columnName + "_proyeccion"

            # Changing columns name with index number
            mapping = {df.columns[0]: col0, df.columns[1]: col1}
            df = df.rename(columns=mapping)
            # Salva la proyeccion normal

            with pd.ExcelWriter(archivo,mode='a', if_sheet_exists="replace") as writer:
                mfila = columnName+"-proyeccion"
                print(mfila)
                df.to_excel(writer, index_label="Periodo",sheet_name=mfila)

            # Calcula el Error por medio del MAPE
            df99 = [df7]
            yy = pd.DataFrame(df99)
            yy = yy.transpose()
            xx = pd.DataFrame(df10[[columnName]])
            print("MAPE " + columnName + " = " + str(mape(xx, yy)))


def agregarAnual(archivo):
    df = []
    hojas = []
    wb2 = load_workbook(archivo)
    sheet_names = wb2.sheetnames
    dt = pd.read_excel(archivo, sheet_name="Sheet1")
    cols = dt.columns
    for c in cols:
        if c != "Periodo":
            hojas.append(c+"-proyeccion")

    for a in hojas:
        if a != "Sheet1":
            col = pd.read_excel(archivo, sheet_name=a)
            print(col)
            col = col.set_index('Periodo')
            df.append(col)
    i = 0

    dt = pd.concat(df, join='inner', axis=1)
    # Salva la proyeccion final de cada serie
    with pd.ExcelWriter(archivo, mode='a', if_sheet_exists="replace") as writer:
        mfila = "Final"
        dt.to_excel(writer, sheet_name=mfila)

    print("Termino")

def proyAnual(archivo):
    df0 = pd.read_excel(archivo, sheet_name="Sheet1")
    dt = pd.read_excel(archivo, sheet_name="Final")
    # Paso 60 adaptado a google sheets
    cols = df0.columns



    df2 = dt.copy(deep=True)

    df3 = df2[df2['Periodo'] > '2020-12-01']

    df3 = df3[df3.columns.drop(list(df3.filter(regex='_ajuste')))]

    df3.columns = df3.columns.str.replace("_proyeccion", "")

    df3 = df3[cols]

    df3 = pd.DataFrame(np.concatenate([df0, df3], axis=0), columns=df0.columns)
    # obtiene categorias, asumiendo Periodo es la primera columna y Total General es la ultima columna
    cats = cols[1:-1]
    tg = cols[-1]
    df3['TotalCat'] = df3[cats].sum(axis=1)
    # Guarda la proyeccion total y la suma de las proyecciones para uso futuro
    totalcat = df3['TotalCat']
    totalgeneral = df3[tg]

    df3 = df3.drop(['TotalCat'], axis=1)

    # Para exportar a excel
    df3sav = df3.copy(deep=True)
    df3sav['Periodo'] = df3sav['Periodo'].astype(str)


    ##archivo = 'P60_Crecimientos.xlsx'
    # df3.to_excel(archivo, index=False)
    df3['Periodo'] = pd.to_datetime(df3['Periodo'])
    df3['Periodo'] = df3.Periodo.dt.year


    df4 = (df3.groupby(df3.Periodo).sum())

    with pd.ExcelWriter(archivo, mode='a', if_sheet_exists="replace") as writer:
        df4.to_excel(writer, sheet_name="Totales")

